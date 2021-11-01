"""
    读取excel表格信息
"""
import openpyxl
import os
import copy
from common.ReadPath import EXCEL_PATH


class ReadExcel:
    def __init__(self):
        self.EXCEL_PATH = EXCEL_PATH
        # 用来存放数据
        self.data = {}
        self.subdata = []
        # 判断文件夹路径是否存在
        if os.path.exists(self.EXCEL_PATH) is False:
            os.makedirs(self.EXCEL_PATH)
            raise(self.EXCEL_PATH+"路径不存在")

    # 读取excel表格路径
    def read_excel_path(self):
        excels_path = []
        for root, dirs, files in os.walk(self.EXCEL_PATH):
            for i in files:
                excel_path = os.path.join(root, i)
                excels_path.append(excel_path)

        for excel_file in excels_path:
            excel_xlsx_file = openpyxl.load_workbook(excel_file, read_only=True)
            for excel_sheet in excel_xlsx_file.sheetnames:
                sheet = excel_xlsx_file[excel_sheet]
                max_row = sheet.max_row
                max_column = sheet.max_column
                if max_row > 1 and max_column > 1:
                    for row in range(2, max_row + 1):
                        for col in range(1, max_column+1):
                            key = sheet.cell(row=1, column=col).value
                            self.data[key] = sheet.cell(row=row, column=col).value
                        data = copy.deepcopy(self.data)
                        self.subdata.append([excel_sheet, data])
            return self.subdata
